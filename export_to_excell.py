import pandas as pd
import pymysql
import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# Подключение к MySQL
connection = pymysql.connect(
    host='127.0.0.1',
    user='root',
    password='',
    database='cardboards_parse',
    charset='utf8mb4',
    cursorclass=pymysql.cursors.DictCursor
)

# SQL-запрос для извлечения данных из таблицы
sql_query = 'SELECT * FROM `mos_gk`'

# Создание объекта курсора
cursor = connection.cursor()

# Выполнение SQL-запроса
cursor.execute(sql_query)

# Извлечение данных
data = cursor.fetchall()

# Закрытие соединения с базой данных
connection.close()

# Запись данных в Excel
wb = Workbook()
ws = wb.active

# Свои имена заголовков
#### для коробок ####
custom_headers = ['id', 'url', 'Уровень 1', 'Уровень 2', 'Уровень 3', 'Уровень 4', 'Уровень 5', 'H1', 'название с делением', 'название с делением/размер', 'Длина', 'Ширина', 'Высота', 'Марка', 'Профиль', 'Цвет', 'Обычная цена']
# Записываем заголовки
ws.append(custom_headers)
# Записываем данные
for row in data:
    row_values = list(row.values())
    # Проверка на соответствие количества значений и заголовков
    if len(row_values) != len(custom_headers):
        print(f"Ошибка: количество значений ({len(row_values)}) не соответствует количеству заголовков ({len(custom_headers)})")
        continue
    for idx, value in enumerate(row_values):
        if custom_headers[idx] == 'url':
            # Создаем гиперссылку
            cell = ws.cell(row=ws.max_row + 1, column=idx + 1, value=value)
            cell.hyperlink = value
            cell.font = Font(color="0563C1", underline="single")
        else:
            ws.cell(row=ws.max_row, column=idx + 1, value=value)

# # Записываем заголовки
# if data:
#     headers = list(data[0].keys())
#     ws.append(headers)

#     # Записываем данные
#     for row in data:
#         ws.append(list(row.values()))

wb.save('outputs/mos_gk_output.xlsx')